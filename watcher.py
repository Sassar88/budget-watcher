#!/usr/bin/env python3
"""Beobachtet eingabe/, extrahiert Buchungen aus Bankauszugs-PDFs/CSVs und schreibt sie in ausgabe/budget.xlsx."""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import shutil
import subprocess
import time
from datetime import datetime
from pathlib import Path

import requests
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

BASE = Path(__file__).resolve().parent
EINGABE = BASE / "eingabe"
VERARBEITET = BASE / "verarbeitet"
FEHLER = BASE / "fehler"
AUSGABE = BASE / "ausgabe"
BACKUPS = AUSGABE / "backups"
BUDGET_XLSX = AUSGABE / "budget.xlsx"
POLL_INTERVAL = 2.0
MAX_BACKUPS = 20

HEADERS = ["Datum", "Empfänger", "Beschreibung", "Betrag (€)", "Kategorie", "Quelle", "Methode"]

SYSTEM_PROMPT = (
    "Du bist ein präziser Parser für Bankauszüge. Extrahiere alle Buchungen aus dem Text und "
    'antworte ausschließlich mit JSON: {"buchungen": [{"datum": "JJJJ-MM-TT", "empfaenger": "", '
    '"beschreibung": "", "betrag": 0.0}]}. '
    "Betrag ist eine Zahl in Euro: Ausgaben negativ, Einnahmen positiv. "
    "Kontostände, Summenzeilen und Kopfzeilen sind keine Buchungen."
)

LEARN_DIR_NAME = "01-gelernt"

CSV_FIELD_ALIASES = {
    "datum": ["buchungstag", "buchungsdatum", "wertstellung", "datum", "value date", "booking date"],
    "empfaenger": [
        "empfänger",
        "empfaenger",
        "beguenstigter/zahlungspflichtiger",
        "begünstigter/zahlungspflichtiger",
        "beguenstigter",
        "begünstigter",
        "zahlungspflichtiger",
        "auftraggeber",
        "name",
        "payee",
    ],
    "beschreibung": ["verwendungszweck", "buchungstext", "beschreibung", "description", "purpose"],
    "betrag": ["betrag", "umsatz", "amount", "value"],
}


def parse_rule_md(text: str) -> dict:
    name = None
    keywords = []
    for line in text.splitlines():
        heading = re.match(r"^#\s+(.+?)\s*$", line)
        if heading and name is None:
            name = heading.group(1)
            continue
        bullet = re.match(r"^\s*[-*]\s+(\S+)", line)
        if bullet:
            keywords.append(bullet.group(1))
    return {"name": name, "keywords": keywords} if name else {}


def parse_settings_md(text: str) -> dict:
    match = re.search(r"default_category\s*:\s*(.+)", text, re.IGNORECASE)
    return {"default_category": match.group(1).strip()} if match else {}


def load_rules(rules_dir: Path) -> dict:
    settings_path = rules_dir / "00-settings.md"
    settings = parse_settings_md(settings_path.read_text(encoding="utf-8")) if settings_path.exists() else {}
    categories = []
    for path in sorted(rules_dir.rglob("*.md"), key=lambda p: p.relative_to(rules_dir).as_posix()):
        if path == settings_path:
            continue
        category = parse_rule_md(path.read_text(encoding="utf-8"))
        if category:
            categories.append(category)
    return {"default_category": settings.get("default_category", "Sonstiges"), "categories": categories}


AI_CFG = yaml.safe_load((BASE / "config" / "ai.yaml").read_text(encoding="utf-8")) or {}
RULES_DIR = BASE / "config" / "rules"


def log(msg: str) -> None:
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}", flush=True)


def unique_path(target: Path) -> Path:
    if not target.exists():
        return target
    for i in range(1, 1000):
        candidate = target.with_name(f"{target.stem}-{i}{target.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Zu viele gleichnamige Dateien für '{target.name}'")


def classify(merchant: str, description: str, rules: dict) -> str:
    text = f" {merchant} {description} ".lower()
    for category in rules.get("categories", []):
        for keyword in category.get("keywords", []):
            if re.search(r"\b" + re.escape(str(keyword).strip().lower()) + r"\b", text):
                return category["name"]
    return str(rules.get("default_category", "Sonstiges"))


def extract_pdf_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def read_text_flexible(path: Path) -> str:
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def detect_csv_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t").delimiter
    except csv.Error:
        return ";"


def match_csv_column(headers: list[str], aliases: list[str]) -> str | None:
    normalized = {h.strip().lower(): h for h in headers}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    for alias in aliases:
        for header_norm, header_orig in normalized.items():
            if alias in header_norm:
                return header_orig
    return None


def extract_csv_rows(csv_path: Path) -> list[dict]:
    raw = read_text_flexible(csv_path)
    delimiter = detect_csv_delimiter(raw[:2000])
    reader = csv.DictReader(io.StringIO(raw), delimiter=delimiter)
    headers = reader.fieldnames or []
    columns = {field: match_csv_column(headers, aliases) for field, aliases in CSV_FIELD_ALIASES.items()}
    missing = [field for field, column in columns.items() if column is None]
    if missing:
        raise ValueError(f"CSV-Spalten nicht erkannt: {', '.join(missing)} (gefunden: {', '.join(headers)})")
    rows = []
    for record in reader:
        rows.append(
            {
                "datum": record.get(columns["datum"], ""),
                "empfaenger": record.get(columns["empfaenger"], ""),
                "beschreibung": record.get(columns["beschreibung"], ""),
                "betrag": record.get(columns["betrag"], ""),
            }
        )
    return rows


def ai_available() -> bool:
    base_url = AI_CFG.get("base_url", "")
    if not base_url:
        return False
    try:
        requests.get(base_url.rstrip("/") + "/models", timeout=2)
        return True
    except requests.RequestException:
        return False


def maybe_start_ai() -> None:
    if not AI_CFG.get("auto_start", False) or ai_available():
        return
    if not shutil.which("ollama"):
        return
    log("Lokale KI wird gestartet (ollama serve) …")
    subprocess.Popen(["ollama", "serve"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    deadline = time.time() + int(AI_CFG.get("startup_timeout_seconds", 90))
    while time.time() < deadline:
        if ai_available():
            log("Lokale KI ist bereit.")
            return
        time.sleep(1)
    log("KI-Start-Timeout – Fallback wird genutzt.")


def call_ai(text: str) -> list[dict]:
    url = AI_CFG["base_url"].rstrip("/") + "/chat/completions"
    payload = {
        "model": AI_CFG.get("model", "local"),
        "temperature": AI_CFG.get("temperature", 0),
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": text},
        ],
    }
    headers = {"Authorization": f"Bearer {AI_CFG.get('api_key', 'local')}"}
    response = requests.post(url, json=payload, headers=headers, timeout=AI_CFG.get("timeout_seconds", 300))
    response.raise_for_status()
    return parse_ai_json(response.json()["choices"][0]["message"]["content"])


def parse_ai_json(content: str) -> list[dict]:
    content = re.sub(r"^\s*```[a-zA-Z]*\s*$|\s*```\s*$", "", content, flags=re.MULTILINE)
    start, end = content.find("{"), content.rfind("}")
    if start == -1 or end == -1:
        raise ValueError("KI-Antwort enthält kein JSON.")
    rows = json.loads(content[start : end + 1]).get("buchungen", [])
    if not isinstance(rows, list):
        raise ValueError("Unerwartetes KI-JSON-Format.")
    return rows


def normalize_date(raw: str) -> str:
    raw = (raw or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%d.%m"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw


def normalize_amount(raw) -> float | None:
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace(" ", "").replace(" ", "").replace("€", "").strip().lstrip("+")
    if re.fullmatch(r"-?\d{1,3}(\.\d{3})*,\d{2}", s):
        s = s.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"-?\d+,\d{1,2}", s):
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


FALLBACK_LINE = re.compile(
    r"^\s*(\d{1,2}\.\d{1,2}\.\d{2,4})\s+"
    r"(?P<mid>.*?)\s*"
    r"(?P<amt>[+-]?\s?\d{1,3}(?:\.\d{3})*,\d{2}|[+-]?\d+(?:[.,]\d{1,2})?)\s*$"
)


def extract_fallback(text: str) -> list[dict]:
    rows = []
    for line in text.splitlines():
        match = FALLBACK_LINE.match(line)
        if not match:
            continue
        middle = match.group("mid").strip()
        parts = middle.split(None, 1)
        rows.append(
            {
                "datum": match.group(1),
                "empfaenger": parts[0] if parts else "",
                "beschreibung": parts[1] if len(parts) > 1 else "",
                "betrag": match.group("amt"),
            }
        )
    return rows


def clean_rows(raw_rows: list[dict]) -> list[dict]:
    rows = []
    for item in raw_rows or []:
        betrag = normalize_amount(item.get("betrag"))
        if betrag is None:
            continue
        rows.append(
            {
                "datum": normalize_date(str(item.get("datum", ""))),
                "empfaenger": str(item.get("empfaenger", "")).strip(),
                "beschreibung": str(item.get("beschreibung", "")).strip(),
                "betrag": betrag,
            }
        )
    return rows


def extract_rows(text: str) -> tuple[list[dict], str]:
    if ai_available():
        try:
            rows = clean_rows(call_ai(text))
            if rows:
                return rows, "KI"
            log("KI lieferte keine verwertbaren Buchungen, versuche Fallback …")
        except Exception as exc:
            log(f"KI-Extraktion fehlgeschlagen ({exc}), versuche Fallback …")
    if AI_CFG.get("fallback", {}).get("enabled", True):
        rows = clean_rows(extract_fallback(text))
        if rows:
            return rows, "Fallback"
    raise ValueError("Keine Buchungen im Auszug erkannt.")


def ensure_workbook() -> Workbook:
    if BUDGET_XLSX.exists():
        return load_workbook(str(BUDGET_XLSX))
    wb = Workbook()
    ws = wb.active
    ws.title = "Buchungen"
    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    for column, width in (("A", 12), ("B", 28), ("C", 34), ("D", 12), ("E", 24), ("F", 28), ("G", 10)):
        ws.column_dimensions[column].width = width
    wb.create_sheet("Übersicht")
    return wb


def backup_workbook() -> None:
    if not BUDGET_XLSX.exists():
        return
    BACKUPS.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copy2(BUDGET_XLSX, BACKUPS / f"budget-{stamp}.xlsx")
    for old in sorted(BACKUPS.glob("budget-*.xlsx"))[:-MAX_BACKUPS]:
        old.unlink()


def existing_signatures(ws) -> set[tuple]:
    signatures = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            betrag = round(float(row[3]), 2)
        except (TypeError, ValueError):
            continue
        signatures.add((str(row[0]), str(row[1] or ""), str(row[2] or ""), betrag))
    return signatures


def rebuild_summary(wb: Workbook) -> None:
    totals: dict[str, list] = {}
    for row in wb["Buchungen"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            betrag = float(row[3])
        except (TypeError, ValueError):
            continue
        category = str(row[4] or "Sonstiges")
        totals.setdefault(category, [0, 0.0])
        if betrag < 0:
            totals[category][0] += 1
            totals[category][1] += betrag
    if "Übersicht" in wb.sheetnames:
        del wb["Übersicht"]
    ws = wb.create_sheet("Übersicht")
    ws.append(["Kategorie", "Anzahl Ausgaben", "Summe Ausgaben (€)"])
    ws.freeze_panes = "A2"
    for column, width in (("A", 30), ("B", 16), ("C", 20)):
        ws.column_dimensions[column].width = width
    for category in sorted(totals, key=lambda name: totals[name][1]):
        count, total = totals[category]
        if count:
            ws.append([category, count, round(total, 2)])
    if any(count for count, _ in totals.values()):
        ws.append(
            [
                "SUMME AUSGABEN",
                sum(count for count, _ in totals.values()),
                round(sum(total for _, total in totals.values()), 2),
            ]
        )


def rebuild_month_comparison(wb: Workbook) -> None:
    data: dict[str, dict[str, float]] = {}
    categories_seen: set[str] = set()
    for row in wb["Buchungen"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            betrag = float(row[3])
        except (TypeError, ValueError):
            continue
        if betrag >= 0:
            continue
        monat = str(row[0])[:7]
        category = str(row[4] or "Sonstiges")
        categories_seen.add(category)
        data.setdefault(monat, {}).setdefault(category, 0.0)
        data[monat][category] += betrag

    if "Monatsvergleich" in wb.sheetnames:
        del wb["Monatsvergleich"]
    ws = wb.create_sheet("Monatsvergleich")
    categories = sorted(categories_seen)
    ws.append(["Monat"] + categories + ["Summe"])
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 12
    for idx in range(2, len(categories) + 3):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    for monat in sorted(data):
        row_values = [round(data[monat].get(category, 0.0), 2) for category in categories]
        ws.append([monat] + row_values + [round(sum(row_values), 2)])


def rebuild_abos(wb: Workbook) -> None:
    grouped: dict[tuple[str, float], set[str]] = {}
    category_by_key: dict[tuple[str, float], str] = {}
    for row in wb["Buchungen"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            betrag = round(float(row[3]), 2)
        except (TypeError, ValueError):
            continue
        empfaenger = str(row[1] or "").strip()
        if betrag >= 0 or not empfaenger:
            continue
        monat = str(row[0])[:7]
        key = (empfaenger.lower(), betrag)
        grouped.setdefault(key, set()).add(monat)
        category_by_key[key] = str(row[4] or "Sonstiges")

    if "Abos" in wb.sheetnames:
        del wb["Abos"]
    ws = wb.create_sheet("Abos")
    ws.append(["Empfänger", "Betrag (€)", "Kategorie", "Anzahl Monate", "Erster Monat", "Letzter Monat"])
    ws.freeze_panes = "A2"
    for column, width in (("A", 28), ("B", 12), ("C", 24), ("D", 14), ("E", 12), ("F", 12)):
        ws.column_dimensions[column].width = width
    entries = [
        (empfaenger, betrag, category_by_key[(empfaenger, betrag)], sorted(monate))
        for (empfaenger, betrag), monate in grouped.items()
        if len(monate) >= 2
    ]
    for empfaenger, betrag, category, monate in sorted(entries, key=lambda e: -len(e[3])):
        ws.append([empfaenger.title(), betrag, category, len(monate), monate[0], monate[-1]])


def significant_keyword(empfaenger: str) -> str | None:
    match = re.match(r"^[A-Za-zÄÖÜäöüß.]{3,}", empfaenger.strip())
    return match.group(0).lower() if match else None


UMLAUT_MAP = str.maketrans({"ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss"})


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower().translate(UMLAUT_MAP)).strip("-")
    return slug or "kategorie"


def learn_from_corrections(rules: dict) -> int:
    """Buchungen, die jemand in budget.xlsx manuell von 'Sonstiges' auf eine echte Kategorie
    umgestellt hat, werden als neue Keyword-Regel unter config/rules/01-gelernt/ gespeichert."""
    if not BUDGET_XLSX.exists():
        return 0
    default_category = rules["default_category"]
    known_categories = {category["name"] for category in rules["categories"]}
    known_keywords = {kw.lower() for category in rules["categories"] for kw in category["keywords"]}

    learned: dict[str, set[str]] = {}
    wb = load_workbook(str(BUDGET_XLSX), read_only=True)
    try:
        for row in wb["Buchungen"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            empfaenger, beschreibung, kategorie = str(row[1] or ""), str(row[2] or ""), str(row[4] or "")
            if kategorie == default_category or kategorie not in known_categories:
                continue
            if classify(empfaenger, beschreibung, rules) != default_category:
                continue
            keyword = significant_keyword(empfaenger)
            if not keyword or keyword in known_keywords:
                continue
            learned.setdefault(kategorie, set()).add(keyword)
    finally:
        wb.close()

    if not learned:
        return 0

    learn_dir = RULES_DIR / LEARN_DIR_NAME
    learn_dir.mkdir(parents=True, exist_ok=True)
    added = 0
    for category, keywords in learned.items():
        path = learn_dir / f"{slugify(category)}.md"
        existing = parse_rule_md(path.read_text(encoding="utf-8")) if path.exists() else {}
        existing_keywords = set(existing.get("keywords", []))
        merged = sorted(existing_keywords | keywords)
        added += len(merged) - len(existing_keywords)
        lines = [f"# {category}", "", "<!-- automatisch vom Wächter aus manuellen Korrekturen ergänzt -->", ""]
        lines += [f"- {kw}" for kw in merged]
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return added


def append_rows(rows: list[dict], source: str, method: str, rules: dict) -> tuple[int, int]:
    backup_workbook()
    wb = ensure_workbook()
    ws = wb["Buchungen"]
    seen = existing_signatures(ws)
    added = skipped = 0
    for row in rows:
        betrag = round(row["betrag"], 2)
        signature = (row["datum"], row["empfaenger"], row["beschreibung"], betrag)
        if signature in seen:
            skipped += 1
            continue
        seen.add(signature)
        ws.append(
            [
                row["datum"],
                row["empfaenger"],
                row["beschreibung"],
                row["betrag"],
                classify(row["empfaenger"], row["beschreibung"], rules),
                source,
                method,
            ]
        )
        added += 1
    rebuild_summary(wb)
    rebuild_month_comparison(wb)
    rebuild_abos(wb)
    AUSGABE.mkdir(parents=True, exist_ok=True)
    wb.save(str(BUDGET_XLSX))
    return added, skipped


def process_file(path: Path) -> None:
    log(f"Verarbeite '{path.name}' …")
    if path.suffix.lower() == ".csv":
        rows = clean_rows(extract_csv_rows(path))
        method = "CSV"
        if not rows:
            raise ValueError("Keine Buchungen in der CSV erkannt.")
    else:
        text = extract_pdf_text(path)
        if not text.strip():
            raise ValueError("Kein Text im PDF gefunden (gescanntes Bild-PDF?)")
        rows, method = extract_rows(text)

    rules = load_rules(RULES_DIR)
    learned = learn_from_corrections(rules)
    if learned:
        log(f"{learned} neue Keyword(s) aus manuellen Korrekturen gelernt.")
        rules = load_rules(RULES_DIR)

    added, skipped = append_rows(rows, path.name, method, rules)
    path.replace(unique_path(VERARBEITET / path.name))
    duplicate_note = f", {skipped} Duplikat(e) übersprungen" if skipped else ""
    log(f"OK: {added} Buchungen ({method}) → budget.xlsx{duplicate_note}, Datei nach verarbeitet/ verschoben.")


def watch_loop(once: bool) -> None:
    for folder in (EINGABE, VERARBEITET, FEHLER, AUSGABE):
        folder.mkdir(parents=True, exist_ok=True)
    log(f"Beobachte {EINGABE} (Strg+C zum Beenden).")
    last_size: dict[str, int] = {}
    while True:
        candidates = sorted(p for p in EINGABE.iterdir() if p.suffix.lower() in (".pdf", ".csv"))
        pending = []
        for path in candidates:
            size = path.stat().st_size
            if last_size.get(path.name) != size:
                last_size[path.name] = size
                pending.append(path)  # Datei ist vermutlich noch in Übertragung
                continue
            last_size.pop(path.name, None)
            try:
                maybe_start_ai()
                process_file(path)
            except Exception as exc:
                log(f"FEHLER bei {path.name}: {exc} – Datei nach fehler/ verschoben.")
                path.replace(unique_path(FEHLER / path.name))
        if once and not pending:
            break
        time.sleep(POLL_INTERVAL)


def main() -> None:
    parser = argparse.ArgumentParser(description="Bankauszug-Wächter für die Budgetplanung")
    parser.add_argument("--once", action="store_true", help="einmalig vorhandene Dateien verarbeiten und beenden")
    args = parser.parse_args()
    try:
        watch_loop(once=args.once)
    except KeyboardInterrupt:
        log("Beendet.")


if __name__ == "__main__":
    main()
